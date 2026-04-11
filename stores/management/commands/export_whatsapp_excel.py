"""
Management command: export_whatsapp_excel
Uso: python manage.py export_whatsapp_excel [--output percorso.xlsx] [--solo-non-contattati]
Uso: python manage.py export_whatsapp_excel --solo-non-contattati

"""

import re
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

from stores.models import Store


# ── Palette colori ──────────────────────────────────────────────────────────
CLR_HEADER_BG  = "1A1A2E"   # sfondo intestazione
CLR_HEADER_FG  = "FFFFFF"   # testo intestazione
CLR_WHATSAPP   = "25D366"   # verde WhatsApp (riga highlight)
CLR_ALT_ROW    = "F2FFF5"   # righe alternate leggero
CLR_BORDER     = "CCCCCC"
CLR_SCORE_HIGH = "22C55E"
CLR_SCORE_MED  = "F59E0B"
CLR_SCORE_LOW  = "EF4444"


def _border(color=CLR_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _extract_number(whatsapp_url: str) -> str:
    """Estrae numero pulito da https://wa.me/+39XXXXXXXXXX"""
    if not whatsapp_url:
        return ""
    m = re.search(r"[\d]{6,15}", whatsapp_url.replace(" ", ""))
    return f"+{m.group(0)}" if m else whatsapp_url


def _score_color(score: int) -> str:
    if score >= 7:
        return CLR_SCORE_HIGH
    if score >= 4:
        return CLR_SCORE_MED
    return CLR_SCORE_LOW


class Command(BaseCommand):
    help = "Esporta numeri WhatsApp degli store in un file Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", "-o",
            default=None,
            help="Percorso file output (default: whatsapp_leads_YYYYMMDD.xlsx)",
        )
        parser.add_argument(
            "--solo-non-contattati",
            action="store_true",
            dest="only_not_contacted",
            help="Esporta solo store non ancora contattati",
        )
        parser.add_argument(
            "--min-score",
            type=int,
            default=0,
            dest="min_score",
            help="Score minimo lead da includere (default: 0 = tutti)",
        )

    def handle(self, *args, **options):
        # ── 1. Query ────────────────────────────────────────────────────────
        qs = Store.objects.exclude(whatsapp_url="").exclude(whatsapp_url__isnull=True)

        if options["only_not_contacted"]:
            qs = qs.exclude(status="contacted").exclude(status="replied").exclude(status="converted")

        # Prendi lead_score dall'analisi più recente (se esiste)
        stores = list(qs.prefetch_related("analyses").order_by("-discovered_at"))

        if options["min_score"] > 0:
            stores = [
                s for s in stores
                if (s.analyses.order_by("-created_at").first().lead_score
                    if s.analyses.exists() else 0) >= options["min_score"]
            ]

        if not stores:
            self.stderr.write("⚠️  Nessuno store con WhatsApp trovato.")
            return

        # ── 2. Workbook ─────────────────────────────────────────────────────
        wb = Workbook()

        self._build_main_sheet(wb, stores)
        self._build_summary_sheet(wb, stores)

        # ── 3. Salva ────────────────────────────────────────────────────────
        fname = options["output"] or f"whatsapp_leads_{datetime.now():%Y%m%d_%H%M}.xlsx"
        wb.save(fname)
        self.stdout.write(self.style.SUCCESS(
            f"✅  Esportati {len(stores)} store → {fname}"
        ))

    # ────────────────────────────────────────────────────────────────────────
    def _build_main_sheet(self, wb: Workbook, stores: list):
        ws = wb.active
        ws.title = "WhatsApp Leads"
        ws.freeze_panes = "A2"

        # Colonne: (label, width)
        columns = [
            ("Brand",            22),
            ("Numero WhatsApp",  20),
            ("Link wa.me",       32),
            ("Email",            30),
            ("Nicchia",          16),
            ("Prodotti",         10),
            ("Lead Score",       12),
            ("Stato",            14),
            ("URL Store",        35),
            ("Scoperto il",      18),
        ]

        # Intestazione
        hdr_font  = Font(bold=True, color=CLR_HEADER_FG, name="Arial", size=11)
        hdr_fill  = PatternFill("solid", fgColor=CLR_HEADER_BG)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (label, width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = _border(CLR_HEADER_BG)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 28

        # Dati
        alt_fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

        for row_idx, store in enumerate(stores, start=2):
            analysis = store.analyses.order_by("-created_at").first()
            score    = analysis.lead_score if analysis else 0
            n_prod   = analysis.product_count if analysis else 0

            numero   = _extract_number(store.whatsapp_url)
            wa_link  = store.whatsapp_url or ""

            row_data = [
                store.name or store.domain,
                numero,
                wa_link,
                store.email or "",
                store.get_niche_display(),
                n_prod,
                score,
                store.get_status_display(),
                store.url,
                store.discovered_at.strftime("%d/%m/%Y") if store.discovered_at else "",
            ]

            is_alt = (row_idx % 2 == 0)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = Font(name="Arial", size=10)
                cell.border = _border()
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                if is_alt:
                    cell.fill = alt_fill

                # Colorazione score
                if col_idx == 7:  # Lead Score
                    cell.font = Font(
                        name="Arial", size=10, bold=True,
                        color=_score_color(score)
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Link cliccabile wa.me
                if col_idx == 3 and wa_link:
                    ws.cell(row=row_idx, column=col_idx).hyperlink = wa_link
                    ws.cell(row=row_idx, column=col_idx).font = Font(
                        name="Arial", size=10,
                        color="0563C1", underline="single"
                    )

                # Link store
                if col_idx == 9:
                    ws.cell(row=row_idx, column=col_idx).hyperlink = store.url
                    ws.cell(row=row_idx, column=col_idx).font = Font(
                        name="Arial", size=10,
                        color="0563C1", underline="single"
                    )

            ws.row_dimensions[row_idx].height = 18

        # Totale in fondo
        last_row = len(stores) + 2
        ws.cell(row=last_row, column=1, value="TOTALE").font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=last_row, column=2, value=f'=COUNTA(B2:B{last_row-1})').font = Font(bold=True, name="Arial")
        ws.cell(row=last_row, column=7, value=f'=AVERAGE(G2:G{last_row-1})').number_format = "0.0"
        ws.cell(row=last_row, column=7).font = Font(bold=True, name="Arial")

        # Filtro automatico
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(stores)+1}"

    # ────────────────────────────────────────────────────────────────────────
    def _build_summary_sheet(self, wb: Workbook, stores: list):
        ws = wb.create_sheet("Riepilogo")

        title_font  = Font(bold=True, size=14, name="Arial", color=CLR_HEADER_BG)
        label_font  = Font(bold=True, size=10, name="Arial")
        value_font  = Font(size=10, name="Arial")
        hdr_fill    = PatternFill("solid", fgColor="E8F5E9")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

        ws["A1"] = "📊 Riepilogo WhatsApp Leads"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")

        ws["A2"] = f"Generato il: {datetime.now():%d/%m/%Y %H:%M}"
        ws["A2"].font = Font(italic=True, size=9, name="Arial", color="666666")
        ws.merge_cells("A2:B2")

        # Statistiche base
        stats = [
            ("Store totali con WhatsApp", len(stores)),
            ("Con email + WhatsApp",
             sum(1 for s in stores if s.email and s.whatsapp_url)),
            ("Già contattati",
             sum(1 for s in stores if s.status in ("contacted", "replied", "converted"))),
            ("Da contattare",
             sum(1 for s in stores if s.status not in ("contacted", "replied", "converted", "rejected", "skip"))),
        ]

        # Score distribution
        analyses = [s.analyses.order_by("-created_at").first() for s in stores]
        scores   = [a.lead_score for a in analyses if a]
        stats += [
            ("Score medio lead", f"{sum(scores)/len(scores):.1f}" if scores else "—"),
            ("Lead score ≥ 7 (alto)",  sum(1 for sc in scores if sc >= 7)),
            ("Lead score 4-6 (medio)", sum(1 for sc in scores if 4 <= sc < 7)),
            ("Lead score < 4 (basso)", sum(1 for sc in scores if sc < 4)),
        ]

        # Nicchie
        from collections import Counter
        niche_counts = Counter(s.get_niche_display() for s in stores)

        row = 4
        ws.cell(row=row, column=1, value="📈 Statistiche").font = Font(bold=True, size=11, name="Arial", color=CLR_HEADER_BG)
        ws.cell(row=row, column=1).fill = hdr_fill
        ws.cell(row=row, column=2).fill = hdr_fill
        row += 1

        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value).font = value_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="🏷️ Per nicchia").font = Font(bold=True, size=11, name="Arial", color=CLR_HEADER_BG)
        ws.cell(row=row, column=1).fill = hdr_fill
        ws.cell(row=row, column=2).fill = hdr_fill
        row += 1

        for nicchia, count in niche_counts.most_common():
            ws.cell(row=row, column=1, value=nicchia).font = label_font
            ws.cell(row=row, column=2, value=count).font = value_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            row += 1